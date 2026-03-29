"""
Utility functions for exams app
"""
from io import BytesIO
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate_grade_sheet_pdf(exam, grades):
    """
    Generate PDF grade sheet for an exam
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f"Grade Sheet - {exam.name}", title_style))

    # Exam details
    detail_style = styles['Normal']
    details = f"""
    <b>Course:</b> {exam.course_offering.course.code} - {exam.course_offering.course.title}<br/>
    <b>Exam Type:</b> {exam.get_exam_type_display()}<br/>
    <b>Date:</b> {exam.date.strftime('%B %d, %Y')}<br/>
    <b>Total Marks:</b> {exam.total_marks}<br/>
    <b>Passing Marks:</b> {exam.passing_marks}<br/>
    <b>Venue:</b> {exam.venue or 'N/A'}<br/>
    <b>Generated:</b> {datetime.now().strftime('%B %d, %Y at %I:%M %p')}
    """
    elements.append(Paragraph(details, detail_style))
    elements.append(Spacer(1, 20))

    # Grades table
    data = [['#', 'Student ID', 'Student Name', 'Marks', 'Grade', 'Status']]

    for idx, grade in enumerate(grades, 1):
        status = 'Pass' if grade.is_passing else 'Fail'
        data.append([
            str(idx),
            grade.student.student_id,
            grade.student.user.get_full_name(),
            f"{grade.marks_obtained}/{exam.total_marks}",
            grade.grade_letter or '-',
            status
        ])

    # Statistics row
    if grades:
        avg_marks = sum(g.marks_obtained for g in grades) / len(grades)
        pass_count = sum(1 for g in grades if g.is_passing)
        pass_rate = (pass_count / len(grades)) * 100

        data.append(['', '', '', '', '', ''])
        data.append(['Statistics', f'Total: {len(grades)}', f'Pass: {pass_count}',
                    f'Fail: {len(grades) - pass_count}', f'Average: {avg_marks:.2f}',
                    f'Pass Rate: {pass_rate:.1f}%'])

    table = Table(data, colWidths=[0.5*inch, 1.2*inch, 2.5*inch, 1*inch, 0.8*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -3), colors.beige),
        ('GRID', (0, 0), (-1, -3), 1, colors.black),
        ('BACKGROUND', (0, -2), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return buffer


def generate_transcript_pdf(transcript):
    """
    Generate PDF transcript for a student
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Header
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=10,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("ERNEST BAI KOROMA UNIVERSITY", title_style))
    elements.append(Paragraph("OFFICIAL TRANSCRIPT", styles['Heading2']))
    elements.append(Spacer(1, 20))

    # Student information
    student = transcript.student
    info_style = styles['Normal']
    student_info = f"""
    <b>Student ID:</b> {student.student_id}<br/>
    <b>Name:</b> {student.user.get_full_name()}<br/>
    <b>Program:</b> {student.program.name}<br/>
    <b>Semester:</b> {transcript.semester}<br/>
    <b>Academic Year:</b> {transcript.academic_year}<br/>
    <b>Issue Date:</b> {transcript.issued_date.strftime('%B %d, %Y') if transcript.issued_date else 'Not Issued'}
    """
    elements.append(Paragraph(student_info, info_style))
    elements.append(Spacer(1, 20))

    # Courses table
    data = [['Course Code', 'Course Title', 'Credits', 'Grade', 'Grade Points']]

    for course in transcript.courses_taken:
        data.append([
            course.get('code', ''),
            course.get('title', ''),
            str(course.get('credits', 0)),
            course.get('grade', ''),
            str(course.get('grade_points', 0))
        ])

    # Summary
    data.append(['', '', '', '', ''])
    data.append(['', 'Total Credits', str(transcript.total_credits), 'GPA', f"{transcript.gpa:.2f}"])
    data.append(['', '', '', 'CGPA', f"{transcript.cgpa:.2f}"])

    table = Table(data, colWidths=[1.2*inch, 3*inch, 0.8*inch, 0.8*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -4), 1, colors.black),
        ('BACKGROUND', (0, -3), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -3), (-1, -1), 'Helvetica-Bold'),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 30))

    # Footer
    footer_text = f"""
    <i>This is an official transcript generated on {datetime.now().strftime('%B %d, %Y')}.</i><br/>
    <i>For verification, contact the Office of the Registrar.</i>
    """
    elements.append(Paragraph(footer_text, info_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_grade_sheet_excel(exam, grades):
    """
    Generate Excel grade sheet for an exam
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Grade Sheet"

    # Header styling
    header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = f"Grade Sheet - {exam.name}"
    ws['A1'].font = Font(bold=True, size=16, color="1e3a8a")
    ws.merge_cells('A1:F1')

    # Exam details
    ws['A3'] = f"Course: {exam.course_offering.course.code} - {exam.course_offering.course.title}"
    ws['A4'] = f"Exam Type: {exam.get_exam_type_display()}"
    ws['A5'] = f"Date: {exam.date.strftime('%B %d, %Y')}"
    ws['A6'] = f"Total Marks: {exam.total_marks}"
    ws['A7'] = f"Passing Marks: {exam.passing_marks}"

    # Column headers
    headers = ['#', 'Student ID', 'Student Name', 'Marks Obtained', 'Grade', 'Status']
    row = 9
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows
    for idx, grade in enumerate(grades, 1):
        row += 1
        status = 'Pass' if grade.is_passing else 'Fail'
        data = [
            idx,
            grade.student.student_id,
            grade.student.user.get_full_name(),
            f"{grade.marks_obtained}/{exam.total_marks}",
            grade.grade_letter or '-',
            status
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col != 3 else 'left')

            # Color code status
            if col == 6:
                if status == 'Pass':
                    cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
                    cell.font = Font(color="155724", bold=True)
                else:
                    cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
                    cell.font = Font(color="721c24", bold=True)

    # Statistics
    if grades:
        avg_marks = sum(g.marks_obtained for g in grades) / len(grades)
        pass_count = sum(1 for g in grades if g.is_passing)
        pass_rate = (pass_count / len(grades)) * 100

        row += 2
        ws[f'A{row}'] = 'Statistics'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f'Total: {len(grades)}'
        ws[f'C{row}'] = f'Pass: {pass_count}'
        ws[f'D{row}'] = f'Fail: {len(grades) - pass_count}'
        ws[f'E{row}'] = f'Average: {avg_marks:.2f}'
        ws[f'F{row}'] = f'Pass Rate: {pass_rate:.1f}%'

    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_promotional_list_excel(promotional_list):
    """
    Generate Excel file for promotional list
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Promotional List"

    # Styling
    header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = f"Promotional List - {promotional_list.program.code} Level {promotional_list.level}"
    ws['A1'].font = Font(bold=True, size=16, color="1e3a8a")
    ws.merge_cells('A1:G1')

    # Details
    ws['A3'] = f"Semester: {promotional_list.semester}"
    ws['A4'] = f"Academic Year: {promotional_list.academic_year}"
    ws['A5'] = f"Generated: {promotional_list.generated_date.strftime('%B %d, %Y')}"

    # Headers
    headers = ['#', 'Student ID', 'Student Name', 'Current Level', 'Next Level', 'CGPA', 'Status']
    row = 7
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data
    promotions = promotional_list.student_promotions.all().order_by('-cgpa')
    for idx, promotion in enumerate(promotions, 1):
        row += 1
        data = [
            idx,
            promotion.student.student_id,
            promotion.student.user.get_full_name(),
            promotion.current_level,
            promotion.next_level or '-',
            f"{promotion.cgpa:.2f}",
            promotion.get_status_display()
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col != 3 else 'left')

            # Color code status
            if col == 7:
                if promotion.status == 'PROMOTED':
                    cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
                elif promotion.status == 'PROBATION':
                    cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
